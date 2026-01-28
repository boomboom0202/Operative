from django.views.generic import ListView
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from operative.models import Resource, Workshop, StatisticsManager

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm


class StatisticsView(ListView):
    template_name = 'table.html'
    context_object_name = 'statistics'

    def get_queryset(self):
        year = int(self.request.GET.get('year', timezone.now().year))
        month = int(self.request.GET.get('month', timezone.now().month))
        today = timezone.now().date()

        data = []

        # Сортировка цехов по ord_s
        for workshop in Workshop.objects.all().order_by('ord_s'):
            rows = []

            # Сортировка ресурсов по priority
            for resource in Resource.objects.all().order_by('priority'):
                month_stats = StatisticsManager.get_monthly_statistics(
                    workshop, resource, year, month
                )

                day_stats = StatisticsManager.get_daily_statistics(
                    workshop, resource, today
                )

                year_stats = StatisticsManager.get_yearly_statistics(
                    workshop, resource, year, current_month=month
                )

                rows.append({
                    'resource': resource,

                    'plan_month': month_stats['plan_month'],
                    'month_plan': month_stats['plan_to_date'],
                    'month_fact': month_stats['fact_month'],
                    'month_diff': month_stats['deviation'],

                    'day_plan': day_stats['plan_daily'],
                    'day_fact': day_stats['fact_daily'],
                    'day_diff': day_stats['deviation'],

                    'year_plan': year_stats['year_plan_to_date'],
                    'year_fact': year_stats['year_fact_to_date'],
                    'year_diff': year_stats['year_deviation'],
                })

            data.append({
                'workshop': workshop,
                'rows': rows
            })

        return data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['year'] = int(self.request.GET.get('year', timezone.now().year))
        context['month'] = int(self.request.GET.get('month', timezone.now().month))
        return context


def export_statistics_excel(request):
    """Экспорт статистики в Excel"""
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    today = timezone.now().date()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Статистика {month:02d}.{year}"

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('A1:R1')
    ws['A1'] = f'Статистика производства за {month:02d}.{year}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    headers = [
        'Код_Цеха', 'Ord_S', 'Цех', 'Цех_KZ', 'Ресурс', 'Ресурс_KZ', 'Priority', 'Единица',
        'План (месяц)', 'План (до даты)', 'Факт (месяц)', 'Откл. (месяц)',
        'План (день)', 'Факт (день)', 'Откл. (день)',
        'План (год)', 'Факт (год)', 'Откл. (год)'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    row_num = 4
    # Сортировка цехов по ord_s
    for workshop in Workshop.objects.all().order_by('ord_s'):
        workshop_start_row = row_num

        # Сортировка ресурсов по priority
        for resource in Resource.objects.all().order_by('priority'):
            month_stats = StatisticsManager.get_monthly_statistics(
                workshop, resource, year, month
            )
            day_stats = StatisticsManager.get_daily_statistics(
                workshop, resource, today
            )
            year_stats = StatisticsManager.get_yearly_statistics(
                workshop, resource, year, current_month=month
            )
            ws.cell(row=row_num, column=1, value=workshop.code)
            ws.cell(row=row_num, column=2, value=workshop.ord_s)
            ws.cell(row=row_num, column=3, value=workshop.name)
            ws.cell(row=row_num, column=4, value=workshop.name_kz)
            ws.cell(row=row_num, column=5, value=resource.name)
            ws.cell(row=row_num, column=6, value=resource.name_kz)
            ws.cell(row=row_num, column=7, value=resource.priority)
            ws.cell(row=row_num, column=8, value=resource.unit)

            ws.cell(row=row_num, column=9, value=month_stats['plan_month'])
            ws.cell(row=row_num, column=10, value=month_stats['plan_to_date'])
            ws.cell(row=row_num, column=11, value=month_stats['fact_month'])
            ws.cell(row=row_num, column=12, value=month_stats['deviation'])

            ws.cell(row=row_num, column=13, value=day_stats['plan_daily'])
            ws.cell(row=row_num, column=14, value=day_stats['fact_daily'])
            ws.cell(row=row_num, column=15, value=day_stats['deviation'])

            ws.cell(row=row_num, column=16, value=year_stats['year_plan_to_date'])
            ws.cell(row=row_num, column=17, value=year_stats['year_fact_to_date'])
            ws.cell(row=row_num, column=18, value=year_stats['year_deviation'])
            
            for col_num in range(1, 19):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if col_num <= 4 else 'right', vertical='center')
                
                if col_num >= 5:
                    cell.number_format = '#,##0.00'

            row_num += 1

        if workshop_start_row < row_num:
            ws.merge_cells(f'A{workshop_start_row}:A{row_num-1}')
            ws.cell(row=workshop_start_row, column=1).alignment = Alignment(
                horizontal='center', vertical='center'
            )

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 12
    for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        ws.column_dimensions[col].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=statistics_{month:02d}_{year}.xlsx'
    wb.save(response)

    return response