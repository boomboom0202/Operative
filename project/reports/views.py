# reports/views.py
from django.shortcuts import render
from django.http import HttpResponse
from .db_connection import get_mssql_connection
from datetime import datetime
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import html
import re

def get_svod_report(date_param):
    try:
        conn = get_mssql_connection()
        cursor = conn.cursor()
        
        cursor.execute("EXEC dbo.rep_cor_svod_new_R @d = ?", date_param)
        
        while cursor.description is None:
            if not cursor.nextset():
                break
        
        results = []
        columns = []
        
        if cursor.description:
            columns = [col[0] for col in cursor.description]
            
            for row in cursor.fetchall():
                row_dict = {}
                for idx, col in enumerate(columns):
                    value = row[idx]
                    if isinstance(value, Decimal):
                        value = float(value)
                    elif hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    elif value is None:
                        value = ''
                    row_dict[col] = value
                results.append(row_dict)
        
        cursor.close()
        conn.close()
        
        excluded = {
            'pr',
            'name',
            'kod_pred',
            's_pred_deport_id',
            'name_kaz',
            'name_rab_kaz',
            'pv_r',
            'pv_v',
            'ord_s',
            's_deport_id',
            'ediz_kaz',
            'grup',
        }
        
        COLUMN_TITLES = {
            'mes': '№',
            'name': 'Наименование предприятия',
            'name_rab': 'Информация',
            'plan_mes': 'План на месяц',
            'ediz': 'Ед. изм.',
            'plan_s': 'План (сутки)',
            'fakt_S': 'Факт (сутки)',
            'delta_s': 'Отклонение (сутки)',
            'plan_m': 'План (месяц)',
            'fakt_m': 'Факт (месяц)',
            'delta_m': 'Отклонение (месяц)',
            'plan_g': 'План (год)',
            'fakt_g': 'Факт (год)',
            'delta_g': 'Отклонение (год)',
            'vipol': 'Выполнение (%)',
        }
        columns = [col for col in columns if col not in excluded]
        if 'plan_mes' in columns and 'ediz' in columns:
            i_plan = columns.index('plan_mes')
            i_ediz = columns.index('ediz')
            columns[i_plan], columns[i_ediz] = columns[i_ediz], columns[i_plan]
        display_columns = [
            {
                'key': col,
                'title': COLUMN_TITLES.get(col, col)
            }
            for col in columns
        ]
        
        return {
            'success': True,
            'columns': display_columns,
            'data': results,
            'count': len(results)
        }
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'data': [],
            'columns': [],
            'count': 0
        }

def group_data_by_enterprise(data):
    MANUAL_SUBGROUPS = {
        "Нурказганский производственный комплекс": [
            "шахта Нурказган",
            "Нурказганская ОФ (технолог.)"
        ],
    }
    grouped = []
    current_enterprise_id = None
    current_subgroup_id = None
    current_enterprise_name = None
    group_index = 0
    subgroup_counters = {}
    
    all_subgroups = set()
    for parent, subgroups in MANUAL_SUBGROUPS.items():
        for subgroup in subgroups:
            all_subgroups.add(subgroup)

    for row in data:
        kod_pred = row.get('kod_pred')
        name = row.get('name', '').strip()

        if kod_pred == 0:
            is_subgroup_of_another = name in all_subgroups
            
            if is_subgroup_of_another:
                if current_enterprise_id not in subgroup_counters:
                    subgroup_counters[current_enterprise_id] = 0
                subgroup_counters[current_enterprise_id] += 1
                
                current_subgroup_id = f"{current_enterprise_id}_sub_{subgroup_counters[current_enterprise_id]}"
                
                grouped.append({
                    'is_header': True,
                    'level': 2,
                    'enterprise_id': current_subgroup_id,
                    'parent_id': current_enterprise_id,
                    'data': row
                })
            else:
                # Это основная группа
                group_index += 1
                current_enterprise_id = f"enterprise_{group_index}"
                current_enterprise_name = name
                current_subgroup_id = None
                
                grouped.append({
                    'is_header': True,
                    'level': 1,
                    'enterprise_id': current_enterprise_id,
                    'parent_id': None,
                    'data': row
                })
        else:
            # Обычный элемент (kod_pred != 0)
            parent = current_subgroup_id if current_subgroup_id else current_enterprise_id
            
            grouped.append({
                'is_header': False,
                'level': 2,
                'parent_id': parent,
                'data': row
            })

    return grouped


def svod_report_page(request):
    date_str = request.GET.get(
        'date',
        datetime.now().strftime('%Y-%m-%d 00:00:00')
    )

    report_data = get_svod_report(date_str)
    
    grouped_data = group_data_by_enterprise(report_data['data']) if report_data['data'] else []

    context = {
        'date': date_str,
        'columns': report_data['columns'],
        'data': report_data['data'],
        'grouped_data': grouped_data,
        'count': report_data['count'],
        'success': report_data['success'],
        'error': report_data.get('error', '')
    }

    return render(request, 'svod.html', context)

def clean_html_entities(text):
        if not isinstance(text, str):
            return text
        text = html.unescape(text)
        text = text.replace('\xa0', ' ')
        text = re.sub(r'\s+', ' ', text)
        text = text.strip()
        return text

def export_svod_excel(request):
    date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d 00:00:00'))
    
    report_data = get_svod_report(date_str)
    
    if not report_data['success'] or not report_data['data']:
        return HttpResponse('Нет данных для экспорта', status=400)
    
    grouped_data = group_data_by_enterprise(report_data['data'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Оперативные показатели"
    
    BLUE_COLOR = '0066B3'      
    BLUE_DARK = '004A85'       
    YELLOW_COLOR = 'FFC627'    
    LIGHT_BLUE = 'E3F2FD'
    LIGHT_BLUE_2 = 'BBDEFB'  # Для подгрупп
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=BLUE_COLOR, end_color=BLUE_COLOR, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    period_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    period_fill = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type='solid')
    
    enterprise_font = Font(name='Arial', size=10, bold=True, color=BLUE_DARK)
    enterprise_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
    
    subgroup_font = Font(name='Arial', size=10, bold=True, color=BLUE_DARK)
    subgroup_fill = PatternFill(start_color=LIGHT_BLUE_2, end_color=LIGHT_BLUE_2, fill_type='solid')
    
    normal_font = Font(name='Arial', size=10)
    
    border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    columns = report_data['columns']
    num_cols = len(columns)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws['A1']
    title_cell.value = 'Оперативные показатели корпорации'
    title_cell.font = Font(name='Arial', size=14, bold=True, color=BLUE_DARK)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color=YELLOW_COLOR, end_color=YELLOW_COLOR, fill_type='solid')
    
    # Дата
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    date_cell = ws['A2']
    date_cell.value = f'Дата: {date_str}'
    date_cell.font = Font(name='Arial', size=10)
    date_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_row_1 = 4
    header_row_2 = 5
    rowspan_columns = ['mes', 'name', 'name_rab', 'ediz', 'plan_mes', 'vipol']
    col_idx = 1
    for col in columns:
        cell_1 = ws.cell(row=header_row_1, column=col_idx)
        cell_1.border = border
        
        if col['key'] in rowspan_columns:
            ws.merge_cells(start_row=header_row_1, start_column=col_idx, 
                          end_row=header_row_2, end_column=col_idx)
            cell_1.value = clean_html_entities(col['title'])
            cell_1.font = header_font
            cell_1.fill = header_fill
            cell_1.alignment = header_alignment
            col_idx += 1
            
        elif col['key'] == 'plan_s':
            ws.merge_cells(start_row=header_row_1, start_column=col_idx, 
                          end_row=header_row_1, end_column=col_idx+2)
            cell_1.value = 'За сутки'
            cell_1.font = period_font
            cell_1.fill = period_fill
            cell_1.alignment = header_alignment
            for i in range(3):
                ws.cell(row=header_row_1, column=col_idx+i).border = border
            
            col_idx += 3
            
        elif col['key'] == 'plan_m':
            ws.merge_cells(start_row=header_row_1, start_column=col_idx, 
                          end_row=header_row_1, end_column=col_idx+2)
            cell_1.value = 'За месяц'
            cell_1.font = period_font
            cell_1.fill = period_fill
            cell_1.alignment = header_alignment
            
            for i in range(3):
                ws.cell(row=header_row_1, column=col_idx+i).border = border
            
            col_idx += 3
            
        elif col['key'] == 'plan_g':
            ws.merge_cells(start_row=header_row_1, start_column=col_idx, 
                          end_row=header_row_1, end_column=col_idx+2)
            cell_1.value = 'За год'
            cell_1.font = period_font
            cell_1.fill = period_fill
            cell_1.alignment = header_alignment
            
            for i in range(3):
                ws.cell(row=header_row_1, column=col_idx+i).border = border
            
            col_idx += 3
    
    col_idx = 1
    for col in columns:
        if col['key'] in rowspan_columns:
            col_idx += 1
            
        elif col['key'] in ['plan_s', 'plan_m', 'plan_g']:
            cell = ws.cell(row=header_row_2, column=col_idx)
            cell.value = 'План'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            col_idx += 1
            
        elif col['key'] in ['fakt_S', 'fakt_m', 'fakt_g']:
            cell = ws.cell(row=header_row_2, column=col_idx)
            cell.value = 'Факт'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            col_idx += 1
            
        elif col['key'] in ['delta_s', 'delta_m', 'delta_g']:
            cell = ws.cell(row=header_row_2, column=col_idx)
            cell.value = 'Откл.'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            col_idx += 1
    
    current_row = 6
    
    for row_data in grouped_data:
        is_header = row_data.get('is_header', False)
        level = row_data.get('level', 1)
        data = row_data['data']
        
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            value = data.get(col['key'], '')
            
            if isinstance(value, (int, float)):
                if col['key'] in ['vipol']:
                    cell.value = value
                    cell.number_format = '0.00'
                else:
                    cell.value = value
                    cell.number_format = '#,##0.00'
            else:
                cleaned_value = clean_html_entities(str(value)) if value else ''
                cell.value = cleaned_value
            
            if is_header:
                if level == 1:
                    cell.font = enterprise_font
                    cell.fill = enterprise_fill
                elif level == 2:
                    cell.font = subgroup_font
                    cell.fill = subgroup_fill
            else:
                cell.font = normal_font
                if col_idx == 3 and row_data.get('parent_id', '').find('_sub_') != -1:  
                    cell.value = '    ' + str(cell.value)
                    
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
        
        current_row += 1
    
    for col_idx in range(1, num_cols + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row_idx in range(4, current_row):
            try:
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'svod_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response